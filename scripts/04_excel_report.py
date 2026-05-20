# ============================================================
# SCRIPT: 04_excel_report.py
# PURPOSE: Generate a formatted, professional Excel report
# WHY: PV teams produce weekly/monthly Excel summaries.
#      Automating this saves hours of manual formatting.
#      openpyxl lets Python write Excel with colors, formulas,
#      charts, and multiple sheets — exactly like doing it manually.
# ============================================================

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ── CONNECT AND LOAD DATA ─────────────────────────────────────

conn = sqlite3.connect("database/pharmacovigilance.db")

# Load all data tables
df_reports = pd.read_sql("SELECT * FROM ae_reports", conn)
df_signals  = pd.read_sql("SELECT * FROM signals WHERE ci_lower > 1 ORDER BY ror_value DESC", conn)
df_monthly  = pd.read_sql("SELECT * FROM monthly_summary ORDER BY year_month", conn)

conn.close()

# ── STYLE DEFINITIONS ─────────────────────────────────────────

# Colors (hex without #)
COLOR_DARK_BLUE   = "1F3864"
COLOR_LIGHT_BLUE  = "2E75B6"
COLOR_ORANGE      = "C55A11"
COLOR_LIGHT_GRAY  = "F2F2F2"
COLOR_RED         = "FF0000"
COLOR_YELLOW      = "FFFF00"
COLOR_GREEN       = "70AD47"

# Header style
def style_header(cell, color=COLOR_DARK_BLUE, font_color="FFFFFF"):
    cell.font = Font(bold=True, color=font_color, size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Side(style='thin', color='FFFFFF')
    cell.border = Border(left=border, right=border, top=border, bottom=border)

# Data cell style
def style_data_cell(cell, bg_color=None):
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    border = Side(style='thin', color='D9D9D9')
    cell.border = Border(left=border, right=border, top=border, bottom=border)

# ── CREATE WORKBOOK ───────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)  # Remove default empty sheet


# ═══════════════════════════════════════════════════════════════
# SHEET 1: SUMMARY DASHBOARD
# ═══════════════════════════════════════════════════════════════

ws_summary = wb.create_sheet("Summary Dashboard")
ws_summary.column_dimensions['A'].width = 30
ws_summary.column_dimensions['B'].width = 20
ws_summary.column_dimensions['C'].width = 20
ws_summary.column_dimensions['D'].width = 20

# Title
ws_summary['A1'] = "PHARMACOVIGILANCE MONITORING REPORT"
ws_summary['A1'].font = Font(bold=True, size=16, color=COLOR_DARK_BLUE)
ws_summary['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
ws_summary['A2'].font = Font(italic=True, size=10, color="808080")

ws_summary.merge_cells('A1:D1')
ws_summary.merge_cells('A2:D2')

# KPI Section Header
ws_summary['A4'] = "KEY PERFORMANCE INDICATORS"
style_header(ws_summary['A4'], COLOR_DARK_BLUE)
ws_summary.merge_cells('A4:D4')

# Calculate KPIs
total_reports   = len(df_reports)
serious_reports = len(df_reports[df_reports['seriousness'] != 'Non-serious'])
fatal_reports   = len(df_reports[df_reports['seriousness'] == 'Death'])
duplicate_rate  = round(df_reports['is_duplicate'].mean() * 100, 1)
signal_count    = len(df_signals[df_signals['signal_strength'].isin(['Strong', 'Moderate'])])

kpis = [
    ("Total AE Reports", total_reports, "", ""),
    ("Serious AE Reports", serious_reports, f"{serious_reports/total_reports*100:.1f}%", "of total"),
    ("Fatal Reports", fatal_reports, f"{fatal_reports/total_reports*100:.1f}%", "of total"),
    ("Duplicate Rate", f"{duplicate_rate}%", "", ""),
    ("Active Signals Detected", signal_count, "", ""),
]

headers = ["Metric", "Value", "Percentage", "Note"]
for col, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=5, column=col, value=h)
    style_header(cell, COLOR_LIGHT_BLUE)

for i, (metric, value, pct, note) in enumerate(kpis, 6):
    ws_summary.cell(row=i, column=1, value=metric).font = Font(bold=True)
    ws_summary.cell(row=i, column=2, value=value).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=i, column=3, value=pct).alignment  = Alignment(horizontal="center")
    ws_summary.cell(row=i, column=4, value=note).font = Font(italic=True, color="808080")
    
    # Alternate row shading
    bg = COLOR_LIGHT_GRAY if i % 2 == 0 else "FFFFFF"
    for col in range(1, 5):
        style_data_cell(ws_summary.cell(row=i, column=col), bg)


# ═══════════════════════════════════════════════════════════════
# SHEET 2: ADVERSE EVENT REPORTS (RAW DATA)
# ═══════════════════════════════════════════════════════════════

ws_reports = wb.create_sheet("AE Reports")

# Select columns to display
display_cols = [
    'report_id', 'report_date', 'drug_name', 'adverse_event',
    'severity', 'seriousness', 'outcome', 'reporter_type',
    'country', 'causality', 'patient_age', 'patient_gender'
]
df_display = df_reports[display_cols].copy()

# Write headers
headers = [col.replace('_', ' ').title() for col in display_cols]
for col_idx, header in enumerate(headers, 1):
    cell = ws_reports.cell(row=1, column=col_idx, value=header)
    style_header(cell)
    # Auto-width
    ws_reports.column_dimensions[get_column_letter(col_idx)].width = max(15, len(header) + 4)

# Write data rows
for row_idx, row_data in df_display.iterrows():
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_reports.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # Color code severity
        if col_idx == 5:   # severity column
            if value == "Fatal":
                cell.fill = PatternFill("solid", fgColor="FF0000")
                cell.font = Font(bold=True, color="FFFFFF")
            elif value == "Life-threatening":
                cell.fill = PatternFill("solid", fgColor="FF6600")
                cell.font = Font(bold=True, color="FFFFFF")
            elif value == "Severe":
                cell.fill = PatternFill("solid", fgColor="FFCC00")
        
        style_data_cell(cell)

# Freeze top row (like "freeze panes" in Excel)
ws_reports.freeze_panes = "A2"

# Auto-filter (dropdown filters on headers)
ws_reports.auto_filter.ref = f"A1:{get_column_letter(len(display_cols))}1"

print("✓ AE Reports sheet created")


# ═══════════════════════════════════════════════════════════════
# SHEET 3: SIGNAL DETECTION RESULTS
# ═══════════════════════════════════════════════════════════════

ws_signals = wb.create_sheet("Signal Detection")

ws_signals['A1'] = "PHARMACOVIGILANCE SIGNAL DETECTION REPORT"
ws_signals['A1'].font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)
ws_signals.merge_cells('A1:G1')

# Explanation
ws_signals['A2'] = "ROR > 1 with 95% CI lower bound > 1 indicates a potential safety signal"
ws_signals['A2'].font = Font(italic=True, size=9, color="808080")
ws_signals.merge_cells('A2:G2')

# Signal headers
sig_headers = ['Drug Name', 'Adverse Event', 'Reports', 'ROR', 'CI Lower', 'CI Upper', 'Signal Strength']
sig_widths   = [20, 25, 10, 10, 10, 10, 18]

for col_idx, (header, width) in enumerate(zip(sig_headers, sig_widths), 1):
    cell = ws_signals.cell(row=3, column=col_idx, value=header)
    style_header(cell, COLOR_ORANGE)
    ws_signals.column_dimensions[get_column_letter(col_idx)].width = width

# Signal data
sig_cols = ['drug_name', 'adverse_event', 'report_count', 'ror_value', 'ci_lower', 'ci_upper', 'signal_strength']
df_sig_display = df_signals[sig_cols].head(50)  # Top 50 signals

for row_idx, row_data in df_sig_display.iterrows():
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_signals.cell(row=row_idx + 4, column=col_idx, value=value)
        
        # Color code signal strength
        if col_idx == 7:
            if value == "Strong":
                cell.fill = PatternFill("solid", fgColor="FF4444")
                cell.font = Font(bold=True, color="FFFFFF")
            elif value == "Moderate":
                cell.fill = PatternFill("solid", fgColor="FF8800")
                cell.font = Font(bold=True, color="FFFFFF")
            elif value == "Weak":
                cell.fill = PatternFill("solid", fgColor="FFEE00")

print("✓ Signal Detection sheet created")


# ═══════════════════════════════════════════════════════════════
# SHEET 4: CHARTS
# ═══════════════════════════════════════════════════════════════

ws_charts = wb.create_sheet("Charts")
ws_charts['A1'] = "PHARMACOVIGILANCE ANALYTICS CHARTS"
ws_charts['A1'].font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)

# Prepare chart data
drug_counts = df_reports['drug_name'].value_counts().head(8).reset_index()
drug_counts.columns = ['Drug', 'Count']

severity_counts = df_reports['severity'].value_counts().reset_index()
severity_counts.columns = ['Severity', 'Count']

# Write drug data for chart
ws_charts['A3'] = "Reports by Drug"
ws_charts['A3'].font = Font(bold=True)
ws_charts['A4'] = "Drug Name"
ws_charts['B4'] = "Report Count"
for i, row in drug_counts.iterrows():
    ws_charts[f'A{i+5}'] = row['Drug']
    ws_charts[f'B{i+5}'] = row['Count']

# Create Bar Chart — Reports by Drug
bar_chart = BarChart()
bar_chart.title = "Adverse Event Reports by Drug"
bar_chart.y_axis.title = "Number of Reports"
bar_chart.x_axis.title = "Drug Name"
bar_chart.style = 10
bar_chart.width = 20
bar_chart.height = 12

data_ref = Reference(ws_charts, min_col=2, min_row=4, max_row=4+len(drug_counts))
cats_ref = Reference(ws_charts, min_col=1, min_row=5, max_row=4+len(drug_counts))
bar_chart.add_data(data_ref, titles_from_data=True)
bar_chart.set_categories(cats_ref)
ws_charts.add_chart(bar_chart, "D3")

# Write severity data for pie chart
ws_charts['A16'] = "Severity Distribution"
ws_charts['A16'].font = Font(bold=True)
ws_charts['A17'] = "Severity"
ws_charts['B17'] = "Count"
for i, row in severity_counts.iterrows():
    ws_charts[f'A{i+18}'] = row['Severity']
    ws_charts[f'B{i+18}'] = row['Count']

# Create Pie Chart — Severity Distribution
pie_chart = PieChart()
pie_chart.title = "AE Severity Distribution"
pie_chart.style = 10
pie_chart.width = 15
pie_chart.height = 12

pie_data = Reference(ws_charts, min_col=2, min_row=17, max_row=17+len(severity_counts))
pie_cats = Reference(ws_charts, min_col=1, min_row=18, max_row=17+len(severity_counts))
pie_chart.add_data(pie_data, titles_from_data=True)
pie_chart.set_categories(pie_cats)
ws_charts.add_chart(pie_chart, "D20")

print("✓ Charts sheet created")


# ── SAVE WORKBOOK ─────────────────────────────────────────────

output_dir = "reports/excel"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    print(f"Created directory: {output_dir}/")

output_path = f"{output_dir}/PV_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(output_path)

print(f"\n✓ Excel report saved: {output_path}")
print("✓ Open it in Excel/LibreOffice to see the formatted report!")